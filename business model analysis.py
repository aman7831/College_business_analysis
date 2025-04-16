
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

# Constants
USD_TO_NPR = 137.54
INFLATION_RATE = 0.0521
YEARS = 5
INTAKES_PER_YEAR = 2
STUDENTS_PER_INTAKE = 15
ADMISSION_FEE = 80000
SEMESTER_FEE = 50000 * 4  # total fee for 4 semesters
EXAM_FEE_USD = 610
TEACHER_COST_PER_SUBJECT = 150000

MARKETING_COST_PER_YEAR = 500000
INFRASTRUCTURE_COST_YEAR_1 = 2000000
ADMIN_COST_PER_STUDENT = 10000
MISC_EXPENSES_PER_YEAR = 300000

exam_fee_npr = EXAM_FEE_USD * USD_TO_NPR

financial_data = []
fee_projection_data = []
cash_flow_data = []

for year in range(1, YEARS + 1):
    inflation_multiplier = (1 + INFLATION_RATE) ** (year - 1)

    admission_fee = ADMISSION_FEE * inflation_multiplier
    semester_fee = SEMESTER_FEE * inflation_multiplier
    teacher_cost = TEACHER_COST_PER_SUBJECT * inflation_multiplier

    total_students = INTAKES_PER_YEAR * STUDENTS_PER_INTAKE
    total_admission_revenue = total_students * admission_fee
    total_semester_revenue = total_students * semester_fee  # already includes 4 semesters per student
    total_exam_revenue = total_students * exam_fee_npr
    total_revenue = total_admission_revenue + total_semester_revenue + total_exam_revenue

    total_subjects_per_year = (5 * 3 + 1)
    total_teacher_cost = total_subjects_per_year * teacher_cost * INTAKES_PER_YEAR

    admin_cost = total_students * ADMIN_COST_PER_STUDENT * inflation_multiplier
    marketing_cost = MARKETING_COST_PER_YEAR * inflation_multiplier
    infra_cost = INFRASTRUCTURE_COST_YEAR_1 if year == 1 else 0
    misc_expense = MISC_EXPENSES_PER_YEAR * inflation_multiplier

    total_expense = total_teacher_cost + admin_cost + marketing_cost + infra_cost + misc_expense
    net_profit = total_revenue - total_expense
    roi = (net_profit / total_expense) * 100 if total_expense > 0 else 0

    financial_data.append({
        "Year": f"Year {year}",
        "Total Students": total_students,
        "Admission Fee / Student": round(admission_fee, 2),
        "Semester Fee / Student": round(semester_fee, 2),
        "Total Revenue": round(total_revenue, 2),
        "Teaching Cost": round(total_teacher_cost, 2),
        "Admin Cost": round(admin_cost, 2),
        "Marketing Cost": round(marketing_cost, 2),
        "Infrastructure Cost": round(infra_cost, 2),
        "Misc Expenses": round(misc_expense, 2),
        "Total Expense": round(total_expense, 2),
        "Net Profit": round(net_profit, 2),
        "ROI (%)": round(roi, 2)
    })

    fee_projection_data.append({
        "Year": f"Year {year}",
        "Admission Fee": round(admission_fee, 2),
        "Semester Fee": round(semester_fee, 2),
        "Exam Fee (NPR, fixed)": round(exam_fee_npr, 2)
    })

    cash_flow_data.append({
        "Year": f"Year {year}",
        "Inflation Rate": INFLATION_RATE,
        "Revenue": round(total_revenue, 2),
        "Expenses": round(total_expense, 2),
        "Net Cash Flow": round(net_profit, 2),
        "Cumulative Cash Flow": None
    })

cumulative = 0
for entry in cash_flow_data:
    cumulative += entry["Net Cash Flow"]
    entry["Cumulative Cash Flow"] = round(cumulative, 2)

# Breakeven Analysis
fixed_cost_per_year = 2000000 + 500000 + (STUDENTS_PER_INTAKE * 2 * ADMIN_COST_PER_STUDENT) + MISC_EXPENSES_PER_YEAR
revenue_per_student = ADMISSION_FEE + (4 * SEMESTER_FEE) + exam_fee_npr
variable_cost_per_student = (TEACHER_COST_PER_SUBJECT * (5 * 3 + 1) * 2) / (STUDENTS_PER_INTAKE * 2)
breakeven_students = fixed_cost_per_year / (revenue_per_student - variable_cost_per_student)
breakeven_data = pd.DataFrame([{
    "Fixed Cost (NPR)": round(fixed_cost_per_year, 2),
    "Revenue per Student": round(revenue_per_student, 2),
    "Variable Cost per Student": round(variable_cost_per_student, 2),
    "Breakeven Number of Students": round(breakeven_students, 2)
}])

# Loan Amortization
loan_amount = 5000000
interest_rate = 0.08
loan_years = 5
emi = loan_amount * interest_rate * (1 + interest_rate) ** loan_years / ((1 + interest_rate) ** loan_years - 1)
loan_schedule = []
remaining = loan_amount

for year in range(1, loan_years + 1):
    interest = remaining * interest_rate
    principal = emi - interest
    remaining -= principal
    loan_schedule.append({
        "Year": f"Year {year}",
        "EMI Payment": round(emi, 2),
        "Principal Paid": round(principal, 2),
        "Interest Paid": round(interest, 2),
        "Remaining Balance": round(remaining, 2)
    })

df_financial = pd.DataFrame(financial_data)
df_fees = pd.DataFrame(fee_projection_data)
df_cashflow = pd.DataFrame(cash_flow_data)
df_loan = pd.DataFrame(loan_schedule)

# Save to Excel
wb = Workbook()
ws1 = wb.active
ws1.title = "Financial Report"
for r in dataframe_to_rows(df_financial, index=False, header=True):
    ws1.append(r)
chart = LineChart()
chart.title = "Revenue vs Profit vs Expense"
chart.y_axis.title = 'Amount (NPR)'
chart.x_axis.title = 'Year'
data = Reference(ws1, min_col=5, min_row=1, max_col=12, max_row=len(df_financial)+1)
chart.add_data(data, titles_from_data=True)
ws1.add_chart(chart, "N2")

ws2 = wb.create_sheet(title="Fee Projection")
for r in dataframe_to_rows(df_fees, index=False, header=True):
    ws2.append(r)

ws3 = wb.create_sheet(title="Cash Flow")
for r in dataframe_to_rows(df_cashflow, index=False, header=True):
    ws3.append(r)

ws4 = wb.create_sheet(title="Breakeven Analysis")
for r in dataframe_to_rows(breakeven_data, index=False, header=True):
    ws4.append(r)

ws5 = wb.create_sheet(title="Loan Amortization")
for r in dataframe_to_rows(df_loan, index=False, header=True):
    ws5.append(r)

wb.save("college_business_model_full_analysis.xlsx")
print("✅ Report generated successfully.")
